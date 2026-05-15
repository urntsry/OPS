import sys
import io
import re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from pathlib import Path

DOWNLOADS = Path(r"c:\Users\888\Downloads")
OUTPUT = Path(r"c:\Users\888\Desktop\python\Project\OPS\docs\seed_hr_data.sql")

lines = []
lines.append("-- HR Data Import from Excel files")
lines.append("-- Generated automatically\n")

def safe_float(v):
    try:
        return float(v) if v else 0
    except:
        return 0

# ============================================================
# 1. OVERTIME - 加班分段表 (26.04月)
# ============================================================
print("Processing overtime file...")
ot_file = DOWNLOADS / "260506加班46hr分段 (26.04月).xlsx"
wb = openpyxl.load_workbook(ot_file, data_only=True)

month_period = "2026-04"
lines.append("\n-- ============ OVERTIME RECORDS (2026-04) ============\n")

record_count = 0
for sheet in wb.worksheets[1:]:
    title = sheet.title.strip()
    
    # Extract employee_id from sheet title like "黎文祥(10235)"
    match = re.search(r'\((\d{5})\)', title)
    if not match:
        # Try to find from row 3 content
        for r in range(1, min(5, sheet.max_row + 1)):
            cell_val = sheet.cell(r, 1).value
            if cell_val:
                m = re.search(r'\((\d{5})\)', str(cell_val))
                if m:
                    match = m
                    break
    
    if not match:
        print(f"  Skipping sheet '{title}' - no employee ID found")
        continue
    
    emp_id = match.group(1)
    
    # Find data rows (skip header rows, data starts after row with '日期')
    data_start = None
    for r in range(1, min(10, sheet.max_row + 1)):
        val = sheet.cell(r, 1).value
        if val and '日期' in str(val):
            data_start = r + 1
            break
    
    if not data_start:
        data_start = 5  # Default
    
    for r in range(data_start, sheet.max_row + 1):
        date_val = sheet.cell(r, 1).value
        weekday_val = sheet.cell(r, 2).value
        ot1_val = sheet.cell(r, 3).value
        ot2_val = sheet.cell(r, 4).value
        note_val = sheet.cell(r, 6).value  # Column 6 has notes like "(8.5改8)"
        
        if not date_val:
            continue
        
        # Parse date: format is "04/01" or "4/1"
        date_str = str(date_val).strip()
        date_match = re.match(r'(\d{1,2})/(\d{1,2})', date_str)
        if not date_match:
            continue
        
        month_num = int(date_match.group(1))
        day_num = int(date_match.group(2))
        record_date = f"2026-{month_num:02d}-{day_num:02d}"
        
        h1 = safe_float(ot1_val)
        h2 = safe_float(ot2_val)
        
        if h1 == 0 and h2 == 0:
            continue
        
        wd = str(weekday_val).strip() if weekday_val else ''
        note_sql = f"'{str(note_val).replace(chr(39), chr(39)+chr(39))}'" if note_val else "NULL"
        
        lines.append(
            f"INSERT INTO overtime_records (profile_id, record_date, weekday, overtime_type1_hours, overtime_type2_hours, note, month_period) "
            f"SELECT id, '{record_date}', '{wd}', {h1}, {h2}, {note_sql}, '{month_period}' "
            f"FROM profiles WHERE employee_id = '{emp_id}' "
            f"ON CONFLICT (profile_id, record_date) DO UPDATE SET overtime_type1_hours = {h1}, overtime_type2_hours = {h2}, weekday = '{wd}', note = {note_sql};"
        )
        record_count += 1

print(f"  Generated {record_count} overtime records")

# ============================================================
# 2. ATTENDANCE - 年度考勤統計表 (114年度 = 2025)
# ============================================================
print("\nProcessing attendance file...")
att_file = DOWNLOADS / "260123年度考勤統計表 (114年度).xlsx"
wb2 = openpyxl.load_workbook(att_file, data_only=True)

att_year = 2025
lines.append(f"\n\n-- ============ LEAVE RECORDS (Year {att_year}) ============\n")

# Process 差假明細 sheet
leave_sheet = wb2['差假明細(2025-01-01~2025-12-31)']
print(f"  Leave sheet rows: {leave_sheet.max_row}")

# Headers at row 1: 部門名稱(0), 員工編號(1), 員工姓名(2), 假別(3), 起迄時間(4), 日數(5), 日數換算(6), 時數(7), total(8), 事由(9)
leave_count = 0
for r in range(2, leave_sheet.max_row + 1):
    emp_id_val = leave_sheet.cell(r, 2).value  # Column B = 員工編號
    if not emp_id_val:
        continue
    emp_id_str = str(emp_id_val).strip()
    if not emp_id_str.isdigit() or len(emp_id_str) != 5:
        continue
    
    leave_type = str(leave_sheet.cell(r, 4).value or '其他').strip()
    time_range = str(leave_sheet.cell(r, 5).value or '').strip()  # "2025-12-31 13:00 ~ 2025-12-31 17:30"
    days_val = safe_float(leave_sheet.cell(r, 6).value)
    hours_val = safe_float(leave_sheet.cell(r, 9).value)  # Column I = total
    reason_val = leave_sheet.cell(r, 10).value  # Column J = 事由
    
    # Parse time range
    if '~' in time_range:
        parts = time_range.split('~')
        start_str = parts[0].strip()
        end_str = parts[1].strip()
    elif '～' in time_range:
        parts = time_range.split('～')
        start_str = parts[0].strip()
        end_str = parts[1].strip()
    else:
        continue
    
    # Ensure proper datetime format
    if len(start_str) <= 10:
        start_str += ' 08:30:00'
    else:
        start_str += ':00'
    if len(end_str) <= 10:
        end_str += ' 17:30:00'
    else:
        end_str += ':00'
    
    leave_type_sql = leave_type.replace("'", "''")
    reason_sql = f"'{str(reason_val).replace(chr(39), chr(39)+chr(39))}'" if reason_val else "NULL"
    
    lines.append(
        f"INSERT INTO leave_records (profile_id, leave_type, start_time, end_time, days, hours, reason, year) "
        f"SELECT id, '{leave_type_sql}', '{start_str}', '{end_str}', {days_val}, {hours_val}, {reason_sql}, {att_year} "
        f"FROM profiles WHERE employee_id = '{emp_id_str}';"
    )
    leave_count += 1

print(f"  Generated {leave_count} leave records")

# Annual leave balance
lines.append(f"\n\n-- ============ ANNUAL LEAVE BALANCE (Year {att_year}) ============\n")
annual_sheet = wb2['年度休假']
print(f"  Annual leave sheet rows: {annual_sheet.max_row}")

# Check header structure
print(f"  Row 1: {[annual_sheet.cell(1, c).value for c in range(1, 22)]}")
print(f"  Row 2: {[annual_sheet.cell(2, c).value for c in range(1, 22)]}")

# Find the actual header row and column positions
header_row = None
for r in range(1, 5):
    for c in range(1, 22):
        val = annual_sheet.cell(r, c).value
        if val and '員工編號' in str(val):
            header_row = r
            break
    if header_row:
        break

if not header_row:
    header_row = 2

# Read headers from this row
headers = [annual_sheet.cell(header_row, c).value for c in range(1, 22)]
print(f"  Headers at row {header_row}: {headers}")

# Find emp_id column
emp_col = None
for c in range(1, 22):
    val = annual_sheet.cell(header_row, c).value
    if val and '員工編號' in str(val):
        emp_col = c
        break

if not emp_col:
    emp_col = 2

# We need to find: 可休, 已休, 轉代金, 保留/遞延, 剩餘
# Let's look at merged headers — scan row for keywords
col_entitled = col_used = col_converted = col_carried = col_remaining = None
for c in range(1, annual_sheet.max_column + 1):
    for r in range(1, 4):
        val = annual_sheet.cell(r, c).value
        if not val:
            continue
        vs = str(val)
        if '可休' in vs or '可使用' in vs:
            col_entitled = c
        elif '已休' in vs and '已休' not in str(annual_sheet.cell(r, c-1).value or ''):
            col_used = c
        elif '代金' in vs:
            col_converted = c
        elif '保留' in vs or '遞延' in vs:
            col_carried = c
        elif '剩餘' in vs or '未休' in vs:
            col_remaining = c

print(f"  Columns: entitled={col_entitled}, used={col_used}, converted={col_converted}, carried={col_carried}, remaining={col_remaining}")

annual_count = 0
for r in range(header_row + 1, annual_sheet.max_row + 1):
    emp_id_val = annual_sheet.cell(r, emp_col).value
    if not emp_id_val:
        continue
    emp_id_str = str(emp_id_val).strip()
    if not emp_id_str.isdigit() or len(emp_id_str) != 5:
        continue
    
    entitled = safe_float(annual_sheet.cell(r, col_entitled).value) if col_entitled else 0
    used = safe_float(annual_sheet.cell(r, col_used).value) if col_used else 0
    converted = safe_float(annual_sheet.cell(r, col_converted).value) if col_converted else 0
    carried = safe_float(annual_sheet.cell(r, col_carried).value) if col_carried else 0
    remaining = safe_float(annual_sheet.cell(r, col_remaining).value) if col_remaining else 0
    
    lines.append(
        f"INSERT INTO annual_leave_balance (profile_id, year, entitled_days, used_days, converted_to_pay, carried_over, remaining) "
        f"SELECT id, {att_year}, {entitled}, {used}, {converted}, {carried}, {remaining} "
        f"FROM profiles WHERE employee_id = '{emp_id_str}' "
        f"ON CONFLICT (profile_id, year) DO UPDATE SET entitled_days = {entitled}, used_days = {used}, converted_to_pay = {converted}, carried_over = {carried}, remaining = {remaining};"
    )
    annual_count += 1

print(f"  Generated {annual_count} annual leave records")

# ============================================================
# 3. BONUS - 盈餘紅利分配表 (115.01+02+03月)
# ============================================================
print("\nProcessing bonus file...")
bonus_file = DOWNLOADS / "260427盈餘紅利分配表 (115.01月+115.02月+115.03月)-Revised 1.xlsx"
wb3 = openpyxl.load_workbook(bonus_file, data_only=True)

quarter_months = ['2026-01', '2026-02', '2026-03']
sheet_names = ['115.01', '115.02', '115.03']

lines.append(f"\n\n-- ============ BONUS MONTHLY (Q1 2026) ============\n")

bonus_count = 0
for month_idx, (ym, sn) in enumerate(zip(quarter_months, sheet_names)):
    if sn not in wb3.sheetnames:
        print(f"  Sheet '{sn}' not found")
        continue
    
    sheet = wb3[sn]
    print(f"  Processing: {sn} -> {ym}, rows: {sheet.max_row}")
    
    # Headers at row 1: [None, '部門名稱', '員工編號', '員工姓名', 'XX月時薪', 'XX月0.5h次數', 'XX月伙食費', 'XX月小計']
    # emp_id at col 3, hourly at col 5, count at col 6, meal at col 7, total at col 8
    
    for r in range(2, sheet.max_row + 1):
        emp_id_val = sheet.cell(r, 3).value  # Column C
        if not emp_id_val:
            continue
        emp_id_str = str(emp_id_val).strip()
        if not emp_id_str.isdigit() or len(emp_id_str) != 5:
            continue
        
        hourly = safe_float(sheet.cell(r, 5).value)
        count = int(safe_float(sheet.cell(r, 6).value))
        meal = safe_float(sheet.cell(r, 7).value)
        total = safe_float(sheet.cell(r, 8).value)
        
        if hourly == 0 and count == 0 and total == 0:
            continue
        
        lines.append(
            f"INSERT INTO bonus_monthly (profile_id, year_month, hourly_rate, half_hour_count, meal_allowance, monthly_total) "
            f"SELECT id, '{ym}', {hourly}, {count}, {meal}, {total} "
            f"FROM profiles WHERE employee_id = '{emp_id_str}' "
            f"ON CONFLICT (profile_id, year_month) DO UPDATE SET hourly_rate = {hourly}, half_hour_count = {count}, meal_allowance = {meal}, monthly_total = {total};"
        )
        bonus_count += 1

print(f"  Generated {bonus_count} bonus monthly records")

# Penalties
lines.append(f"\n\n-- ============ BONUS PENALTIES (Q1 2026) ============\n")
penalty_sheet = wb3['獎懲結算']
print(f"  Penalties sheet rows: {penalty_sheet.max_row}")
print(f"  Row 1: {[penalty_sheet.cell(1, c).value for c in range(1, 9)]}")
print(f"  Row 2: {[penalty_sheet.cell(2, c).value for c in range(1, 9)]}")

# Headers: [None, '部門名稱', '員工編號', '員工姓名', '獎懲事由', '獎懲', '備註', None]
penalty_count = 0
for r in range(2, penalty_sheet.max_row + 1):
    emp_id_val = penalty_sheet.cell(r, 3).value  # Column C
    if not emp_id_val:
        continue
    emp_id_str = str(emp_id_val).strip()
    if not emp_id_str.isdigit() or len(emp_id_str) != 5:
        continue
    
    reason = str(penalty_sheet.cell(r, 5).value or '').strip()
    pen_type = str(penalty_sheet.cell(r, 6).value or '其他').strip()
    note = penalty_sheet.cell(r, 7).value
    
    if not reason:
        continue
    
    # The penalty type cell might contain the deduction info
    # Try to extract amount from penalty type or a separate column
    amount = 0
    # Check if there's a numeric value in the penalty type
    amount_match = re.search(r'(\d+)', pen_type)
    if amount_match:
        amount = float(amount_match.group(1))
    
    reason_sql = reason.replace("'", "''")
    pen_type_sql = pen_type.replace("'", "''")
    note_sql = f"'{str(note).replace(chr(39), chr(39)+chr(39))}'" if note else "NULL"
    
    lines.append(
        f"INSERT INTO bonus_penalties (profile_id, year_month, reason, penalty_type, amount, note) "
        f"SELECT id, '2026-01', '{reason_sql}', '{pen_type_sql}', {amount}, {note_sql} "
        f"FROM profiles WHERE employee_id = '{emp_id_str}';"
    )
    penalty_count += 1

print(f"  Generated {penalty_count} penalty records")

# Write output
with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

total_inserts = len([l for l in lines if l.startswith('INSERT')])
print(f"\n=== DONE ===")
print(f"Output: {OUTPUT}")
print(f"Total SQL statements: {total_inserts}")
