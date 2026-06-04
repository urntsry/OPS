import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

# Simulate the compare logic with real files
wb_att = openpyxl.load_workbook(r'c:\Users\888\Downloads\115年.04月出勤明細 (原始未算加班版).xlsx', data_only=True)
wb_wh = openpyxl.load_workbook(r'c:\Users\888\Downloads\115.04月工時統計單.xlsx', data_only=True)

ws_att = wb_att['Sheet1']

# Parse attendance
att_data = []
for row in ws_att.iter_rows(min_row=4, max_row=ws_att.max_row, max_col=22, values_only=True):
    r = list(row)
    if not r[1] or not str(r[1]).strip():
        continue
    emp_id = str(r[1]).strip()
    if len(emp_id) != 5 or not emp_id.isdigit():
        continue
    att_data.append({
        'emp_id': emp_id,
        'name': str(r[2] or ''),
        'dept': str(r[0] or ''),
        'date': str(r[3] or ''),
        'weekday': str(r[4] or ''),
        'holiday': str(r[5] or ''),
        'clock_in': str(r[9]) if r[9] else None,
        'clock_out': str(r[10]) if r[10] else None,
        'anomaly': str(r[14] or ''),
    })

print(f'Attendance records loaded: {len(att_data)}')

# Parse work hours
wh_data = {}
for sn in wb_wh.sheetnames:
    ws = wb_wh[sn]
    emp_id = None
    m = sn.find('(')
    if m >= 0:
        emp_id = sn[m+1:m+6]
    if not emp_id or len(emp_id) != 5:
        continue
    if emp_id not in wh_data:
        wh_data[emp_id] = {}
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=16, values_only=True):
        r = list(row)
        if not r[0] or '合計' in str(r[0]):
            continue
        date_key = str(r[0]).strip()
        start_time = str(r[2]).strip() if r[2] else ''
        if not start_time:
            continue
        if date_key not in wh_data[emp_id] or start_time < wh_data[emp_id][date_key]:
            wh_data[emp_id][date_key] = start_time

print(f'Work hours employees loaded: {len(wh_data)}')
print()

# Run compare
results = {'pass': [], 'exclude': [], 'review': []}

for att in att_data:
    if not att['clock_in'] and not att['clock_out']:
        continue
    
    date_full = att['date']
    if '-' in date_full:
        date_short = f"{date_full[5:7]}/{date_full[8:10]}"
    else:
        date_short = date_full
    
    emp_id = att['emp_id']
    qr_start = wh_data.get(emp_id, {}).get(date_short, None)
    
    has_early = att['clock_in'] and att['clock_in'] < '08:00'
    has_evening = att['clock_out'] and att['clock_out'] > '18:00'
    
    if not has_early and not has_evening:
        continue
    
    early_ot = 0
    early_status = 'exclude'
    
    if has_early:
        if qr_start and qr_start < '08:00':
            early_ot = 0.5
            early_status = 'pass'
        elif qr_start and qr_start >= '08:00' and qr_start < '08:30':
            early_ot = 0.5
            early_status = 'review'
        elif qr_start and qr_start >= '08:30':
            early_ot = 0
            early_status = 'exclude'
        else:
            # No QR data: if no evening OT either, auto-exclude
            if not has_evening:
                early_ot = 0
                early_status = 'exclude'
            else:
                early_ot = 0
                early_status = 'review'
    
    evening_ot = 0
    if has_evening:
        h = int(att['clock_out'][:2])
        m = int(att['clock_out'][3:5])
        total_min = (h - 18) * 60 + m
        evening_ot = round(total_min / 30) * 0.5
    
    if has_early and has_evening:
        final_status = early_status
    elif has_early:
        final_status = early_status
    else:
        final_status = 'pass'
    
    results[final_status].append({
        'emp': emp_id,
        'name': att['name'],
        'date': date_short,
        'weekday': att['weekday'],
        'clock_in': att['clock_in'],
        'clock_out': att['clock_out'],
        'qr': qr_start,
        'early': early_ot if early_status != 'exclude' else 0,
        'evening': evening_ot,
    })

print('========== COMPARE RESULTS ==========')
print(f'  PASS (auto-approved):    {len(results["pass"])} records')
print(f'  EXCLUDE (auto-removed):  {len(results["exclude"])} records')
print(f'  REVIEW (needs confirm):  {len(results["review"])} records')
print()

print('--- PASS examples (first 10) ---')
for r in results['pass'][:10]:
    print(f'  {r["emp"]} {r["name"]:6s} {r["date"]} {r["weekday"]} in:{r["clock_in"]} out:{r["clock_out"]} QR:{r["qr"]} early:{r["early"]} evening:{r["evening"]}')
print()

print('--- EXCLUDE examples (first 10) ---')
for r in results['exclude'][:10]:
    print(f'  {r["emp"]} {r["name"]:6s} {r["date"]} {r["weekday"]} in:{r["clock_in"]} out:{r["clock_out"]} QR:{r["qr"]}')
print()

print('--- REVIEW (ALL - needs confirmation) ---')
for r in results['review']:
    print(f'  {r["emp"]} {r["name"]:6s} {r["date"]} {r["weekday"]} in:{r["clock_in"]} out:{r["clock_out"]} QR:{r["qr"]} early:{r["early"]} evening:{r["evening"]}')
print()

# Summary
emp_totals = {}
for r in results['pass']:
    if r['emp'] not in emp_totals:
        emp_totals[r['emp']] = {'name': r['name'], 'early': 0, 'evening': 0, 'total': 0}
    emp_totals[r['emp']]['early'] += r['early']
    emp_totals[r['emp']]['evening'] += r['evening']
    emp_totals[r['emp']]['total'] += r['early'] + r['evening']

print(f'--- EMPLOYEE SUMMARY (top 25 by total OT) ---')
overtime_46 = 0
for emp_id, t in sorted(emp_totals.items(), key=lambda x: -x[1]['total'])[:25]:
    flag = ' <<< OVER 46hr' if t['total'] > 46 else ''
    if t['total'] > 46:
        overtime_46 += 1
    print(f'  {emp_id} {t["name"]:8s} early:{t["early"]:5.1f} evening:{t["evening"]:5.1f} TOTAL:{t["total"]:6.1f}{flag}')

print(f'\n  Total employees with OT: {len(emp_totals)}')
print(f'  Employees over 46hr: {overtime_46}')

# Cross-check with known data: 50028 should have early_ot on 4/13, 4/14, 4/15
print('\n--- VERIFICATION: 50028 賴葦蓁 ---')
for r in results['pass'] + results['review'] + results['exclude']:
    if r['emp'] == '50028':
        status = 'PASS' if r in results['pass'] else 'REVIEW' if r in results['review'] else 'EXCLUDE'
        print(f'  [{status}] {r["date"]} {r["weekday"]} in:{r["clock_in"]} out:{r["clock_out"]} QR:{r["qr"]} early:{r["early"]} evening:{r["evening"]}')
